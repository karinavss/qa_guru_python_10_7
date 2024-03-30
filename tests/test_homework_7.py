import csv
import os
from pypdf import PdfReader
from openpyxl import load_workbook


def test_csv_file(create_archive):
    path = os.path.join(create_archive, 'csv_file.csv')
    with open(path) as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        names = [row['hey'] for row in reader]
    assert 'hello' in names


def test_xlsx_file(create_archive):
    path = os.path.join(create_archive, 'file-xlsx.xlsx')
    open_xlsx = load_workbook(path)
    sheet = open_xlsx.active
    name = sheet.cell(row=3, column=2).value
    assert name == 'Mara'

def test_pdf_page_content(create_archive):
    path = f'{create_archive}/Python Testing with pytest.pdf'
    text_to_search = "The source code for the Cards project"
    with open(path, 'rb') as file:
        reader = PdfReader(file)
        text_found = any(text_to_search in page.extract_text() for page in reader.pages)
        assert text_found, f"Текст \"{text_to_search}\" не найден в PDF файле."